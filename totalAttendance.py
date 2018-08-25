#! python3
# -*- coding: utf-8 -*-
"""
Created on Mon Aug 13 14:39:40 2018

@author: btrev

Accepts Churchbuilder attendance spreadsheets, calculates the total attendance
for each individual over the given period and builds a new spreadsheet from
this data.
"""

import pyexcel, openpyxl, os, sys, re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def file_input():
    """User inputs spreadsheet filename and corresponding group name."""
    filename = input('Please enter a spreadsheet filename e.g. attendance.ods '
                     '(or type q to quit)\n')
    if filename == 'q' or filename == 'Q':
        return 1
    if not os.path.isfile(filename):
        raise Exception('Error: Must be a valid filename in the current '
                        'working directory.')    
    if filename[-4:] != '.ods':
        raise Exception('Error: File must have the extension .ods')    
    group = input('Please enter the group name for this register. \n')
    return [filename,group]
   
def convert_to_xlsx(file):
    """Convert the given .ods file into a .xlsx with the given group name in
    the new filename."""
    filename, group = file
    #remove any illegal characters from group for use in filename
    legalGroup = ''.join(c for c in group if c.isalnum())
    xlsxName = 'attendance ' + legalGroup + '.xlsx'
    array = pyexcel.get_array(file_name = filename)
    pyexcel.save_as(array = array, dest_file_name = xlsxName)
    return [xlsxName, group]

def check_dates(file):
    """Checks that the top row of the spreadsheet contains dates only."""
    filename, group = file
    sheet = openpyxl.load_workbook(filename).active
    #regex will return value if there is a date
    dateRegex = re.compile(r'^\d?\d/\d\d$')
    for column in range(2, sheet.max_column):
            date = sheet.cell(1,column).value
            datecheck = dateRegex.search(date)
            if datecheck == None:
                raise Exception('Incorrect spreadsheet format: row 1 must be'
                                'dates only.')
    return

def get_names(attendance, file):
    """Builds a dictionary with all the names that appear on the spreadsheet,
    values are dictionaries containing group name, to store attendance later."""
    filename, group = file
    sheet = openpyxl.load_workbook(filename).active
    #regexes to check column 1 contains names only
    nameRegex = re.compile(r'^[a-zA-z\-]+ [a-zA-z\-]+$')
    for row in range(2, sheet.max_row):        #Cell A1 is empty
            name = sheet['A' + str(row)].value
            namecheck = nameRegex.search(name)
            if namecheck == None:
                raise Exception('Incorrect spreadsheet format: column 1 must ' 
                                'be names only.')
            else:
                attendance.setdefault(name, {})    
                attendance[name].setdefault(group,0)
    return

def sum_attendance_data(attendance, file):
    """Sums the 'Y's that appear on the spreadsheet to calculate total
    attendance and stores the data in the attendance dictionary."""
    filename, group = file
    sheet = openpyxl.load_workbook(filename).active
    #check the body of the spreadsheet consists only of Y and empty cells
    max_column_letter = get_column_letter(sheet.max_column)
    finalCell = max_column_letter + str(sheet.max_row)
    for row in sheet['A2':finalCell]:
        name = row[0].value
        for cell in row[1:]:
            if cell.value != 'Y' and cell.value != None:
                raise Exception('Incorrect spreadsheet format: data must '
                                "consist of 'Y's and empty cells only.")
            #count the attendance for each name
            if cell.value == 'Y':
                attendance[name][group] += 1
    return

def gather_attendance_data():
    """Allows multiple spreadsheets to be input by the user for different 
    groups. The attendance data for all the groups is stored in the attendance
    dictionary."""
    attendance = {}
    groups = []
    while True:
        try:
            file = file_input()
            if file == 1:
                break
            groups.append(file[1])  #save group name for later
            xlsxFile = convert_to_xlsx(file)
            check_dates(xlsxFile)
            get_names(attendance, xlsxFile)
            sum_attendance_data(attendance, xlsxFile)
            os.remove(xlsxFile[0])
        except Exception as err:
            print(err)
            #delete any new .xlsx files
            if 'xlsxFile' in locals():
                if os.path.isfile(xlsxFile[0]):
                    os.remove(xlsxFile[0])
            continue
    
    return [attendance, groups]

def write_totals_sheet(attendance, groups):
    """Writes the total attendance data to a new spreadsheet. Column 1 is names
    row 1 is group names."""  
    if attendance == {}:
        return
    wb = openpyxl.Workbook()
    totalsSheet = wb.active
    #group names as column headers
    for i in range(len(groups)):
        totalsSheet.cell(1,i+2).font = Font(name = 'Calibri', bold = True)
        totalsSheet.cell(1,i+2).value = groups[i]
        #set column width
        col = get_column_letter(i+2)
        totalsSheet.column_dimensions[col].width = len(groups[i])
        
    #width of name column same as longest name
    totalsSheet.column_dimensions['A'].width = len(max(attendance, key = len))
    #enter data row by row
    row = 2
    column = 1
    for name, groupAtt in sorted(attendance.items()):
        totalsSheet.cell(row,1).value = name
        for group, attVal in groupAtt.items():
            column = groups.index(group) + 2    #find correct group column
            totalsSheet.cell(row,column).value = attVal
        row = row + 1
    
    wb.save('Total Attendance.xlsx')
            
def main():
    attendance, groups = gather_attendance_data()
    write_totals_sheet(attendance,groups)
    return 0

if __name__ == '__main__':
    sys.exit(main())
