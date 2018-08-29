#! python3
# -*- coding: utf-8 -*-
"""
Created on Mon Aug 13 14:39:40 2018

@author: btrev

Accepts Churchbuilder attendance spreadsheets, calculates the total attendance
for each individual over the given period and builds a new spreadsheet from
this data.
"""

import pyexcel, openpyxl, os, sys, re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def Max_row(sheet):
    """
    A new definition for the max_row method for openpyxl worksheet objects.
    """
    for max_row, row in enumerate(sheet,1):
        if all(c.value is None for c in row):
            max_row -= 1
            break
    
    return max_row

def register_input(registers):
    """User inputs spreadsheet filename and corresponding group name."""
    while True:
        filename = input('Please enter a spreadsheet filename e.g. attendance.ods '
                         '(or type q to quit)\n')
        if filename == 'q' or filename == 'Q':
            return 1
        if not os.path.isfile(filename):
            raise Exception('Error: Must be a valid filename in the current '
                            'working directory.')    
        if filename[-4:] != '.ods':
            raise Exception('Error: File must have the extension .ods')  
        if filename in registers.values():
            raise Exception('Error: This file has already been chosen.')
            
        group = input('Please enter the group name for this register. \n')
        if group in registers:
            raise Exception("Error: There is already a group named '{0}', "
                            'please choose a different name'.format(group))
        registers[group] = filename
    
    return 0
   
def convert_to_xlsx(registers):
    """
    Convert the given .ods files into .xlsx with the given group name in
    the new filename. Updates the registers dict with new .xlsx filenames
    """
    for group, filename in registers.items():
        #remove any illegal characters from group for use in filename
        legalGroup = ''.join(c for c in group if c.isalnum())
        xlsxName = 'attendance ' + legalGroup + '.xlsx'
        array = pyexcel.get_array(file_name = filename)
        pyexcel.save_as(array = array, dest_file_name = xlsxName)
        registers[group] = xlsxName
    
    return 

def check_dates(registers):
    """Checks that the top row of the spreadsheet contains dates only."""
    #regex will return value if there is a date
    dateRegex = re.compile(r'^\d?\d/\d\d$')
    
    for filename in registers.values():
        sheet = openpyxl.load_workbook(filename).active
        for column in range(2, sheet.max_column):
            date = sheet.cell(1,column).value
            datecheck = dateRegex.search(date)
            if datecheck == None:
                raise Exception("Incorrect spreadsheet format for '{0}': row 1"
                                ' must be dates only.'.format(filename))
    return

def check_names(registers):
    """Checks that column 1 of the spreadsheet contains names only."""
    #regex will return a value if there is a name
    nameRegex = re.compile(r'^[a-zA-z\-]+ [a-zA-z\-]+$')
    
    for filename in registers.values():
        sheet = openpyxl.load_workbook(filename).active
        for row in range(2, Max_row(sheet) + 1):        #Cell A1 is empty
            name = sheet['A' + str(row)].value
            namecheck = nameRegex.search(name)
            if namecheck == None:
                raise Exception("Incorrect spreadsheet format for '{0}': colum"
                                'n 1 must be names only.'.format(filename))
    return

def check_Ys(registers):
    """
    Checks that the body of the spreadsheet contains 'Y's and empty cells only.
    """
    for filename in registers.values():
        sheet = openpyxl.load_workbook(filename).active
        #find reference to bottom right corner cell
        max_column_letter = get_column_letter(sheet.max_column)
        finalCell = max_column_letter + str(Max_row(sheet))
        for row in sheet['A2':finalCell]:
            for cell in row[1:]:
                if cell.value != 'Y' and cell.value != None:
                    raise Exception("Incorrect spreadsheet format for '{0}': "
                                    "data must consist of 'Y's and empty cells"
                                    'only.'.format(filename))
    return

def get_names(registers):
    """
    Builds a dictionary with all the names that appear on the spreadsheet,
    values are dicts containing group name, to store attendance later.
    """
    attendance = {}
    
    for group, filename in registers.items():
        sheet = openpyxl.load_workbook(filename).active
        for row in range(2, Max_row(sheet) + 1):        #Cell A1 is empty
            name = sheet['A' + str(row)].value
            attendance.setdefault(name, {})    
            attendance[name].setdefault(group,0)
    
    return attendance

def sum_attendance_data(attendance, registers):
    """
    Sums the 'Y's that appear on the spreadsheet to calculate total
    attendance and stores the data in the attendance dictionary.
    """
    for group, filename in registers.items():
        print(group)
        sheet = openpyxl.load_workbook(filename).active
        #find reference to bottom right corner cell
        max_column_letter = get_column_letter(sheet.max_column)
        finalCell = max_column_letter + str(Max_row(sheet))
        for row in sheet['A2':finalCell]:
            name = row[0].value
            for cell in row[1:]:
                #count the attendance for each name
                if cell.value == 'Y':
                    print(name, ':', attendance[name])
                    attendance[name][group] += 1
    return

def gather_attendance_data():
    """
    The attendance data for all the registers is summed and stored in the 
    'attendance' dictionary.
    """
    registers = {}
    while True:
        try:
            q = register_input(registers)
            if q == 1 and registers != {}:
                convert_to_xlsx(registers)
                check_dates(registers)
                check_names(registers)
                check_Ys(registers)
                attendance = get_names(registers)
                sum_attendance_data(attendance, registers)
                #return a list of group names for the write function
                groups = list(registers.keys())
                for xlsxFile in registers.values():
                    os.remove(xlsxFile)
                break
            #if there is no data return empty dict and list
            elif q == 1 and registers == {}:
                return ({},[])
        except Exception as err:
            print(err)
            #delete any new .xlsx files
            for filename in registers.values():
                if os.path.isfile(filename) and filename[-5:] == '.xlsx':
                    os.remove(filename)
    
    return (attendance, groups)

def write_totals_sheet(attendance, groups):
    """
    Writes the total attendance data to a new spreadsheet. Column 1 is names
    row 1 is group names.
    """  
    #if there is no data, do nothing
    if attendance == {}:
        return
    wb = openpyxl.Workbook()
    totalsSheet = wb.active
    #group names as column headers
    for i, group in enumerate(groups):
        totalsSheet.cell(1,i+2).font = Font(name = 'Calibri', bold = True)
        totalsSheet.cell(1,i+2).value = group
        #set column width
        col = get_column_letter(i+2)
        totalsSheet.column_dimensions[col].width = max(3, len(group) + 1)
        
    #width of name column same as longest name
    totalsSheet.column_dimensions['A'].width = len(max(attendance, key = len))
    #enter data row by row
    row = 2
    for name, groupAtt in sorted(attendance.items()):
        totalsSheet.cell(row,1).value = name
        for group, attVal in groupAtt.items():
            column = groups.index(group) + 2    #find correct group column
            totalsSheet.cell(row,column).value = attVal
        row = row + 1
    
    wb.save('Total Attendance.xlsx')
            
def main():
    attendance, groups = gather_attendance_data()
    write_totals_sheet(attendance,groups)
    return 0

if __name__ == '__main__':
    sys.exit(main())
